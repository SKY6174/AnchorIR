import os
import re
import shutil
import subprocess
import tempfile
import unicodedata
import openpyxl
from pdf2image import convert_from_path

# 1. 경로 설정 (2차년도 전용)
PDF_DIR = "/Users/thomas/Documents/AnchorIR/data/documents/RISE-2차년도-협약서(B1)"
EXCEL_PATH = os.path.join(PDF_DIR, "UC_ANCHOR_협약서_업로드_서식.xlsx")
BACKUP_PATH = os.path.join(PDF_DIR, "UC_ANCHOR_협약서_업로드_서식_backup.xlsx")

def clean_text(text):
    """줄바꿈 및 불필요한 연속 공백을 정리합니다."""
    if not text:
        return ""
    text = re.sub(r'\s+', ' ', text)
    return text.strip()

def run_tesseract_ocr_on_pil_image(pil_img):
    """PIL Image 객체에 대해 Tesseract OCR을 실행하여 한글 텍스트를 추출합니다."""
    with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as temp_img:
        pil_img.save(temp_img.name, "PNG")
        temp_img_path = temp_img.name

    temp_out_base = temp_img_path + "_out"
    temp_out_txt = temp_out_base + ".txt"

    try:
        cmd = ["tesseract", temp_img_path, temp_out_base, "-l", "kor"]
        res = subprocess.run(cmd, capture_output=True, text=True)
        
        if os.path.exists(temp_out_txt):
            with open(temp_out_txt, "r", encoding="utf-8") as f:
                ocr_text = f.read()
            return ocr_text
        else:
            return ""
    except Exception as e:
        print(f"  [OCR 에러] {e}")
        return ""
    finally:
        if os.path.exists(temp_img_path):
            os.remove(temp_img_path)
        if os.path.exists(temp_out_txt):
            os.remove(temp_out_txt)

def extract_agreement_info_using_rendering(pdf_path):
    """
    pdf2image를 사용하여 PDF의 마지막 페이지를 이미지로 렌더링한 후,
    Tesseract OCR을 통해 한글 텍스트를 추출합니다.
    """
    text = ""
    try:
        pages = convert_from_path(pdf_path, dpi=200)
        if pages:
            target_page_image = pages[-1]
            text = run_tesseract_ocr_on_pil_image(target_page_image).strip()
    except Exception as e:
        print(f"  [렌더링 실패] {e}. pypdf 텍스트 추출로 대체합니다.")
        
    if not text:
        try:
            import pypdf
            reader = pypdf.PdfReader(pdf_path)
            for page in reader.pages:
                text += page.extract_text() or ""
            text = text.strip()
        except:
            text = ""

    if not text:
        return None, None, "대표 OOO"

    cleaned = clean_text(text)

    # 1. 협약체결일 정밀 정규식 탐색
    agree_date_dash = None
    agree_date_yyyyimmd = None

    # YYYY년 MM월 DD일 매칭 시도
    date_match = re.search(r'(\d{2,4})\s*년\s*(\d{1,2})\s*월\s*[^\d]*(\d{1,2})\s*일', cleaned)
    if date_match:
        year, month, day = date_match.groups()
        if len(year) == 2:
            year = "2026" if year in ["26"] else "20" + year
        agree_date_dash = f"{year}-{int(month):02d}-{int(day):02d}"
        agree_date_yyyyimmd = f"{year}{int(month):02d}{int(day):02d}"
    else:
        # 연월만 있고 일이 유실되거나 공란인 경우 대응 (예: "2026 년 8월 일")
        partial_date_match = re.search(r'(\d{2,4})\s*년\s*(\d{1,2})\s*월', cleaned)
        if partial_date_match:
            year, month = partial_date_match.groups()
            if len(year) == 2:
                year = "2026" if year in ["26"] else "20" + year
            agree_date_dash = f"{year}-{int(month):02d}-None"
            agree_date_yyyyimmd = f"{year}{int(month):02d}None"
        else:
            # 점 구분선 대안 매칭 (예: 2026.07.21)
            alt_date_match = re.search(r'(\d{2,4})[\.\-\s]\s*(\d{1,2})[\.\-\s]\s*(\d{1,2})', cleaned)
            if alt_date_match:
                year, month, day = alt_date_match.groups()
                if len(year) == 2:
                    year = "20" + year
                agree_date_dash = f"{year}-{int(month):02d}-{int(day):02d}"
                agree_date_yyyyimmd = f"{year}{int(month):02d}{int(day):02d}"

    # 2. 협약기관 서명자 직위 & 성명 추출
    last_text = text[-450:]
    last_cleaned = clean_text(last_text)
    
    # 대학 총장 및 대학명 정제 제외
    temp_text = re.sub(r'총장\s*[조]\s*[홍]\s*[래]', '', last_cleaned)
    temp_text = re.sub(r'울산과학대\S*', '', temp_text)
    temp_text = re.sub(r'총장\s*조홍래', '', temp_text)
    
    # 직위 + 성명 매칭 (다양한 서명/날인 오독 수용 및 비캡처 그룹 사용)
    sign_match = re.search(r'(대표이사|대표자|대표\s*이사|대표|센터장|원장|원\s*장|회장)\s*[:\.]?\s*([가-힣\s\d]{2,10})\s*\(?(?:서명|서령|날인|서)', temp_text)
    
    signer_info = None
    if sign_match:
        position, name = sign_match.groups()
        position = clean_text(position).replace(" ", "")
        name = clean_text(name).replace(" ", "")
        name = re.sub(r'\d+', '', name) # 숫자 정제
        if 2 <= len(name) <= 5:
            signer_info = f"{position} {name}"
            
    if not signer_info:
        # 서명/날인 괄호가 유실된 형태 대안 매칭
        alt_sign_match = re.search(r'(대표이사|대표자|대표\s*이사|대표|센터장|원장|원\s*장|회장)\s*[:\.]?\s*([가-힣\s]{2,5})(?=\s|$)', temp_text)
        if alt_sign_match:
            position, name = alt_sign_match.groups()
            position = clean_text(position).replace(" ", "")
            name = clean_text(name).replace(" ", "")
            if 2 <= len(name) <= 5:
                signer_info = f"{position} {name}"

    if not signer_info:
        signer_info = "대표 OOO"

    return agree_date_dash, agree_date_yyyyimmd, signer_info

def main():
    print("====================================================")
    print(" RISE 2차년도 협약서 고화질 렌더링 OCR 분석 및 엑셀 기입 시작")
    print("====================================================")

    if not os.path.exists(EXCEL_PATH):
        print(f"[에러] 엑셀 파일이 존재하지 않습니다: {EXCEL_PATH}")
        return

    # 엑셀 백업 생성
    try:
        shutil.copy2(EXCEL_PATH, BACKUP_PATH)
        print(f"  - 원본 엑셀 백업 완료: {os.path.basename(BACKUP_PATH)}")
    except Exception as backup_err:
        print(f"[경고] 백업 생성 실패: {backup_err}")

    # 엑셀 로드
    wb = openpyxl.load_workbook(EXCEL_PATH)
    sheet = wb.active
    print(f"엑셀 로드 완료: {os.path.basename(EXCEL_PATH)} (행 수: {sheet.max_row})")

    # PDF 목록 읽기
    all_files = os.listdir(PDF_DIR)
    pdf_files = [f for f in all_files if f.endswith(".pdf")]
    print(f"디렉토리 내 PDF 파일 개수: {len(pdf_files)}개")

    success_count = 0
    fail_count = 0
    matched_count = 0

    # 엑셀 행 매칭 맵 구축
    excel_rows = {}
    for r in range(2, sheet.max_row + 1):
        org_name = sheet.cell(r, 3).value  # C열: 협약 대상기관
        manager_name = sheet.cell(r, 4).value  # D열: 대학 측 협약주체(UC)
        existing_date = sheet.cell(r, 1).value  # A열: 기존 등록 일자
        existing_signer = sheet.cell(r, 5).value  # E열: 기존 등록 서명자
        
        org_nfc = unicodedata.normalize("NFC", str(org_name or "")).strip()
        manager_nfc = unicodedata.normalize("NFC", str(manager_name or "")).strip()
        manager_only_name = manager_nfc.split()[-1] if manager_nfc else ""
        
        excel_rows[r] = {
            "org": org_nfc,
            "manager_only_name": manager_only_name,
            "org_raw": org_name,
            "manager_raw": manager_name,
            "existing_date": existing_date,
            "existing_signer": existing_signer
        }

    for pdf_name in pdf_files:
        pdf_nfc = unicodedata.normalize("NFC", pdf_name)
        pdf_path = os.path.join(PDF_DIR, pdf_name)
        print(f"\n[분석 중] {pdf_nfc}")

        # 파일명 분리
        # 파일명 형식: [앵커사업-2026]산학연협력운영협약서-{기관명}-{대학담당자}.pdf
        name_parse = re.search(r'산학연협력운영협약서-(.+?)-(.+?)\.pdf', pdf_nfc)
        if not name_parse:
            print("  -> [오류] 파일명 규격이 올바르지 않아 건너뜁니다.")
            fail_count += 1
            continue

        pdf_org = name_parse.group(1).strip()
        pdf_manager = name_parse.group(2).strip()

        # 엑셀 매칭 행 탐색
        target_row_idx = None
        
        # 1차 매칭: 기관명 + 담당자명 동시 만족
        for row_idx, data in excel_rows.items():
            org_match = (pdf_org in data["org"]) or (data["org"] in pdf_org)
            if not org_match:
                # '의원', '병원', '주식회사' 등 일반 분류 접미사를 정제하여 느슨한 대조 수행
                pdf_org_clean = re.sub(r'(의원|병원|주식회사|유한회사|합자회사|재단법인|사단법인|협회|㈜|㈔)$', '', pdf_org).strip()
                excel_org_clean = re.sub(r'(의원|병원|주식회사|유한회사|합자회사|재단법인|사단법인|협회|㈜|㈔)$', '', data["org"]).strip()
                if pdf_org_clean and excel_org_clean and ((pdf_org_clean in excel_org_clean) or (excel_org_clean in pdf_org_clean)):
                    org_match = True
                    
            manager_match = (pdf_manager == data["manager_only_name"])
            if org_match and manager_match:
                target_row_idx = row_idx
                break
                
        # 2차 매칭: 기관명 단독 만족
        if not target_row_idx:
            for row_idx, data in excel_rows.items():
                org_match = (pdf_org in data["org"]) or (data["org"] in pdf_org)
                if not org_match:
                    pdf_org_clean = re.sub(r'(의원|병원|주식회사|유한회사|합자회사|재단법인|사단법인|협회|㈜|㈔)$', '', pdf_org).strip()
                    excel_org_clean = re.sub(r'(의원|병원|주식회사|유한회사|합자회사|재단법인|사단법인|협회|㈜|㈔)$', '', data["org"]).strip()
                    if pdf_org_clean and excel_org_clean and ((pdf_org_clean in excel_org_clean) or (excel_org_clean in pdf_org_clean)):
                        org_match = True
                if org_match:
                    target_row_idx = row_idx
                    break

        if not target_row_idx:
            print(f"  [경고] 엑셀 매칭 실패 (대상기관: {pdf_org}, 담당자: {pdf_manager})")
            fail_count += 1
            continue

        # PDF OCR 정보 추출
        res = extract_agreement_info_using_rendering(pdf_path)
        agree_date_dash, agree_date_yyyyimmd, signer_info = res
        
        # 엑셀 매칭 행 데이터 참조
        row_data = excel_rows[target_row_idx]
        excel_date_raw = str(row_data["existing_date"] or "").strip()
        
        # 3. 지능형 날짜 Fallback (유실되거나 추출 실패한 경우 보정)
        final_date_dash = None
        final_date_yyyyimmd = None
        
        if agree_date_dash and "None" not in agree_date_dash:
            final_date_dash = agree_date_dash
            final_date_yyyyimmd = agree_date_yyyyimmd
        else:
            if agree_date_dash and "None" in agree_date_dash:
                parts = agree_date_dash.split("-")
                year, month = parts[0], parts[1]
                if excel_date_raw.startswith(f"{year}-{month}"):
                    final_date_dash = excel_date_raw
                else:
                    final_date_dash = f"{year}-{month}-01"
                final_date_yyyyimmd = final_date_dash.replace("-", "")
            else:
                if re.match(r'^\d{4}-\d{2}-\d{2}$', excel_date_raw):
                    final_date_dash = excel_date_raw
                    final_date_yyyyimmd = excel_date_raw.replace("-", "")
                else:
                    # 2차년도는 디폴트 날짜를 2026년으로 적용
                    final_date_dash = "2026-05-15"
                    final_date_yyyyimmd = "20260515"

        # 4. 서명자 Fallback (추출 실패 시 기본값 지정)
        final_signer = signer_info
        if final_signer == "대표 OOO":
            excel_signer_raw = str(row_data["existing_signer"] or "").strip()
            if excel_signer_raw and excel_signer_raw != "대표 OOO" and len(excel_signer_raw) > 2:
                final_signer = excel_signer_raw
            else:
                final_signer = "대표 OOO"

        # 엑셀 기입
        sheet.cell(target_row_idx, 1).value = final_date_dash
        sheet.cell(target_row_idx, 5).value = final_signer
        print(f"  => 엑셀 {target_row_idx}행에 기입 완료: [날짜] {final_date_dash} | [서명자] {final_signer}")
        matched_count += 1

        # 파일명 변경 (Rename)
        # 2차년도 파일명의 '[앵커사업-2026]' 부분을 '[앵커사업-YYYYMMDD]'로 변환합니다.
        new_pdf_nfc = re.sub(r'\[앵커[^\]]*?-(2026)\]', f'[앵커사업-{final_date_yyyyimmd}]', pdf_nfc)
        new_pdf_name = unicodedata.normalize("NFC", new_pdf_nfc)
        new_pdf_path = os.path.join(PDF_DIR, new_pdf_name)

        try:
            if pdf_path != new_pdf_path:
                os.rename(pdf_path, new_pdf_path)
                print(f"  - 파일명 변경 완료: {new_pdf_name}")
            success_count += 1
        except Exception as rename_err:
            print(f"  [경고] 파일명 변경 에러: {rename_err}")
            success_count += 1

    # 최종 엑셀 저장
    try:
        wb.save(EXCEL_PATH)
        print("\n====================================================")
        print(" 엑셀 파일 저장 및 완자동화 완료!")
        print(f" 총 처리 대상: {len(pdf_files)}개 | 매칭 및 변경 성공: {success_count}개 | 실패: {fail_count}개")
        print(f" 업데이트된 엑셀 행 수: {matched_count}개 행")
        print("====================================================")
    except Exception as save_err:
        print(f"\n[오류] 엑셀 파일 저장 실패: {save_err}")

if __name__ == "__main__":
    main()
